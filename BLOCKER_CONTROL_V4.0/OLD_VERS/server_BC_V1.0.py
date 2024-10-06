import socket
import threading
import os
import xml.etree.ElementTree as ET
from openpyxl import Workbook, load_workbook
import time

# Configuração do servidor
HOST = '127.0.0.1'
PORT = 65535
clients = {}
client_identifiers = {}
locations = {}
items = {}
deliveries = {}
client_count = 0
gerente_socket = None  # Socket do cliente gerente

# Criação dos arquivos Excel para itens enviados e entregues
excel_file = 'itens_enviados.xlsx'
entregues_file = 'tabela_de_itens_entregados.xlsx'


# Função para salvar informações no Excel
def save_to_excel(item_name, location_id, quantity):
    if not os.path.exists(excel_file):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Item", "Local", "Quantidade"])
    else:
        workbook = load_workbook(excel_file)
        sheet = workbook.active

    sheet.append([item_name, locations[location_id], quantity])
    workbook.save(excel_file)


# Função para mover itens para o arquivo de itens entregues
def move_to_delivered(item_name):
    if os.path.exists(excel_file):
        wb_sent = load_workbook(excel_file)
        sheet_sent = wb_sent.active
        data = None

        for row in sheet_sent.iter_rows(min_row=2, values_only=True):
            if row[0] == item_name:
                data = row
                break

        if data:
            # Remove do arquivo de itens enviados
            for row in sheet_sent.iter_rows(min_row=2):
                if row[0].value == item_name:
                    sheet_sent.delete_rows(row[0].row)
                    break
            wb_sent.save(excel_file)

            # Adiciona no arquivo de itens entregues
            if not os.path.exists(entregues_file):
                wb_delivered = Workbook()
                sheet_delivered = wb_delivered.active
                sheet_delivered.append(["Item", "Local", "Quantidade"])
            else:
                wb_delivered = load_workbook(entregues_file)
                sheet_delivered = wb_delivered.active

            sheet_delivered.append(data)
            wb_delivered.save(entregues_file)


# Função para gerar XML de confirmação de envio
def generate_xml_confirmation(item_name, location_id, quantity):
    root = ET.Element("ItemConfirmation")
    ET.SubElement(root, "Item").text = item_name
    ET.SubElement(root, "Location").text = locations[location_id]
    ET.SubElement(root, "Quantity").text = str(quantity)
    ET.SubElement(root, "Status").text = "Sent"

    tree = ET.ElementTree(root)
    xml_filename = f"{item_name}_confirmation.xml"
    tree.write(xml_filename)
    print(f"Arquivo XML gerado: {xml_filename}")


# Função para salvar locais no arquivo locais.txt
def save_location_to_txt(location_id, location_name):
    with open('locais.txt', 'a') as f:
        f.write(f"{location_id},{location_name}\n")


# Função para carregar os locais do arquivo locais.txt
def load_locations_from_txt():
    if os.path.exists('locais.txt'):
        with open('locais.txt', 'r') as f:
            for line in f.readlines():
                location_id, location_name = line.strip().split(',')
                locations[location_id] = location_name


def alert_gerente(mensagem):
    if gerente_socket:
        gerente_socket.send(f"ALERTA: {mensagem}\n".encode())


def handle_client(conn, addr, is_gerente=False):
    global gerente_socket
    if is_gerente:
        gerente_socket = conn  # Armazena o socket do gerente
        print(f'Gerente conectado: {addr}')
        conn.send("Conectado como gerente.\n".encode())
        while True:
            time.sleep(3600)  # Envia a mensagem de alerta a cada uma hora
            alert_gerente("Você está recebendo este alerta a cada 1 hora até que o erro seja resolvido.")
    else:
        print(f'CLIENTE conectado: {addr}')
        conn.send("Conectado ao servidor!\n".encode())

        while True:
            try:
                data = conn.recv(1024).decode()
                if not data:
                    break

                command = data.split('|')[0].upper()

                if command == 'SEND ITEM':
                    try:
                        item_name = data.split('|')[1]
                        location_id = data.split('|')[2]
                        quantity = int(data.split('|')[3])

                        if location_id in locations:
                            generate_xml_confirmation(item_name, location_id, quantity)
                            save_to_excel(item_name, location_id, quantity)
                            conn.send(f"Item {item_name} enviado para {locations[location_id]}.\n".encode())
                        else:
                            alert_gerente(f"Item {item_name} foi enviado para o local errado.")
                            conn.send(f"Erro: Item {item_name} está no local errado e não pode ser entregue.\n".encode())
                            conn.close()
                            break
                    except (KeyError, IndexError, ValueError):
                        conn.send("Erro ao enviar item. Verifique os parâmetros.\n".encode())

                elif command == 'CONFIRM DELIVERY':
                    try:
                        item_name = data.split('|')[1]
                        location_id = data.split('|')[2]
                        if item_name in deliveries and deliveries[item_name] == location_id:
                            conn.send(f"Entrega do item {item_name} confirmada no local {locations[location_id]}.\n".encode())
                            move_to_delivered(item_name)
                        else:
                            raise KeyError("Item ou local inválido.")
                    except KeyError:
                        conn.send(f"Erro: Item ou local inválido para confirmação de entrega.\n".encode())
                    except IndexError:
                        conn.send("Erro: Formato incorreto para confirmar entrega.\n".encode())

                else:
                    conn.send("Comando inválido.\n".encode())

            except Exception as e:
                conn.send(f"Erro na operação: {e}\n".encode())
                break

        conn.close()
        print(f'Cliente desconectado: {addr}')


def start_server():
    server = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    server.bind((HOST, PORT))
    server.listen(5)
    print(f'Servidor iniciado em {HOST}:{PORT}...')

    while True:
        conn, addr = server.accept()
        # Verifica se é o cliente gerente
        is_gerente = addr[1] == 65534  # Porta diferente usada para o gerente
        client_thread = threading.Thread(target=handle_client, args=(conn, addr, is_gerente))
        client_thread.start()


if __name__ == "__main__":
    load_locations_from_txt()  # Carrega os locais salvos no arquivo locais.txt
    start_server()